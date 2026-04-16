from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class FixtureSummary:
    path: Path
    sheet: str | None
    row_count: int
    columns: list[str]
    sample_rows: list[dict[str, Any]]


def load_fixture_summaries(root: Path) -> list[FixtureSummary]:
    old_data = Path(root) / "Old Data"
    summaries: list[FixtureSummary] = []
    for path in sorted(old_data.glob("*.csv")):
        summaries.append(_summarize_csv(path))
    for path in sorted(old_data.glob("*.xlsx")):
        summaries.extend(_summarize_xlsx(path))
    return summaries


def load_haru_2026_results(root: Path) -> list[dict[str, Any]]:
    path = Path(root) / "Old Data" / "haru_2026_makuuchi_all_days_results.xlsx"
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read old Excel fixtures") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Results"] if "Results" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    columns = [str(cell or "").strip() for cell in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {columns[i]: row[i] if i < len(row) else None for i in range(len(columns))}
        if any(value is not None and value != "" for value in record.values()):
            out.append(record)
    return out


def write_fixture_summary_csv(summaries: list[FixtureSummary], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["path", "sheet", "row_count", "columns", "sample_rows"],
        )
        writer.writeheader()
        for summary in summaries:
            writer.writerow(
                {
                    "path": str(summary.path),
                    "sheet": summary.sheet or "",
                    "row_count": summary.row_count,
                    "columns": " | ".join(summary.columns),
                    "sample_rows": repr(summary.sample_rows[:3]),
                }
            )


def _summarize_csv(path: Path) -> FixtureSummary:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
    return FixtureSummary(
        path=path,
        sheet=None,
        row_count=len(rows),
        columns=list(reader.fieldnames or []),
        sample_rows=rows[:5],
    )


def _summarize_xlsx(path: Path) -> list[FixtureSummary]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read old Excel fixtures") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    summaries: list[FixtureSummary] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            summaries.append(FixtureSummary(path, ws.title, 0, [], []))
            continue
        columns = [str(cell or "").strip() for cell in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {columns[i]: row[i] if i < len(row) else None for i in range(len(columns))}
            if any(value is not None and value != "" for value in record.values()):
                records.append(record)
        summaries.append(FixtureSummary(path, ws.title, len(records), columns, records[:5]))
    return summaries
